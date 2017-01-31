import pytz
import calendar
import collections
from functools import partial

# NOTE: Python 3 compatibility
try:
    from urlparse import urlparse, parse_qs
except ImportError:
    from urllib.parse import urlparse, parse_qs
from datetime import datetime, timedelta

from django.conf import settings
from django.core.mail import EmailMessage
from django.core.management.base import BaseCommand, CommandError
from django.core.validators import URLValidator, EmailValidator
from django.utils import timezone
from django.utils.dateparse import parse_datetime

from openpyxl import Workbook

from seed_services_client import (HubApiClient, IdentityStoreApiClient,
                                  StageBasedMessagingApiClient,
                                  MessageSenderApiClient)


def mk_validator(django_validator):
    def validator(inputstr):
        django_validator()(inputstr)
        return inputstr
    return validator


def midnight(timestamp):
    return timestamp.replace(hour=0, minute=0, second=0, microsecond=0)


def one_month_after(timestamp):
    weekday, number_of_days = calendar.monthrange(
        timestamp.year, timestamp.month)
    return timestamp + timedelta(days=number_of_days)


def midnight_validator(inputstr):
    return midnight(datetime.strptime(inputstr, '%Y-%m-%d')).replace(
        tzinfo=pytz.timezone(settings.TIME_ZONE))


class ExportSheet(object):

    def __init__(self, sheet, headers=None):
        self._sheet = sheet
        self.set_header(headers or [])

    def set_header(self, headers):
        self._headers = headers
        for index, header in enumerate(headers):
            self._sheet.cell(row=1, column=index + 1, value=header)

    def get_header(self):
        return self._headers

    def add_row(self, row):
        row_number = self._sheet.max_row + 1
        for key, value in row.items():
            if isinstance(key, int):
                col_idx = key
            else:
                col_idx = self._headers.index(key) + 1

            cell = self._sheet.cell(
                row=row_number,
                column=col_idx)
            cell.value = value


class ExportWorkbook(object):

    def __init__(self):
        self._workbook = Workbook()

    def add_sheet(self, sheetname, position):
        return ExportSheet(self._workbook.create_sheet(sheetname, position))

    def save(self, file_name):
        return self._workbook.save(file_name)


def parse_cursor_params(cursor):
    parse_result = urlparse(cursor)
    params = parse_qs(parse_result.query)
    return dict([(key, value[0]) for key, value in params.items()])


class Command(BaseCommand):

    workbook_class = ExportWorkbook

    help = ('Generate an XLS spreadsheet report on registrations '
            'and write it to disk')

    def add_arguments(self, parser):
        parser.add_argument(
            '--start',
            type=midnight_validator, default=midnight(timezone.now()),
            help=('The start of the reporting range (YYYY-MM-DD). '
                  'Defaults to today according to the configured timezone.'))
        parser.add_argument(
            '--end',
            type=midnight_validator,
            default=None,
            help=('The end of the reporting range (YYYY-MM-DD). '
                  'Defaults to exactly 1 month after `--start`'))
        parser.add_argument(
            '--output-file', type=str, default=None,
            help='The file to write the report to.'
        )
        parser.add_argument(
            '--email-to', type=mk_validator(EmailValidator),
            default=[], action='append',
            help='Who to email the report to.'
        )
        parser.add_argument(
            '--email-from', type=mk_validator(EmailValidator),
            default=settings.DEFAULT_FROM_EMAIL,
            help='Which address to send the email from',
        )
        parser.add_argument(
            '--email-subject', type=str,
            default='Seed Control Interface Generated Report',
            help='The subject of the email',
        )
        parser.add_argument(
            '--hub-url', type=mk_validator(URLValidator),
            default=settings.HUB_URL,
        )
        parser.add_argument(
            '--hub-token', type=str, default=settings.HUB_TOKEN
        )
        parser.add_argument(
            '--identity-store-url', type=mk_validator(URLValidator),
            default=settings.IDENTITY_STORE_URL)
        parser.add_argument(
            '--identity-store-token', type=str,
            default=settings.IDENTITY_STORE_TOKEN)
        parser.add_argument(
            '--sbm-url', type=mk_validator(URLValidator))
        parser.add_argument(
            '--sbm-token', type=str)
        parser.add_argument(
            '--ms-url', type=mk_validator(URLValidator))
        parser.add_argument(
            '--ms-token', type=str)

    def handle(self, *args, **kwargs):
        self.identity_cache = {}
        self.messageset_cache = {}
        hub_token = kwargs['hub_token']
        hub_url = kwargs['hub_url']
        id_store_token = kwargs['identity_store_token']
        id_store_url = kwargs['identity_store_url']
        sbm_token = kwargs['sbm_token']
        sbm_url = kwargs['sbm_url']
        ms_token = kwargs['ms_token']
        ms_url = kwargs['ms_url']
        start_date = kwargs['start']
        end_date = kwargs['end']
        output_file = kwargs['output_file']

        email_recipients = kwargs['email_to']
        email_sender = kwargs['email_from']
        email_subject = kwargs['email_subject']

        if not sbm_url:
            raise CommandError(
                'Please make sure the --sbm-url is set.')

        if not sbm_token:
            raise CommandError(
                'Please make sure the --sbm-token is set.')

        if not ms_url:
            raise CommandError(
                'Please make sure the --ms-url is set.')

        if not ms_token:
            raise CommandError(
                'Please make sure the --ms-token is set.')

        if not output_file:
            raise CommandError(
                'Please specify --output-file.')

        if end_date is None:
            end_date = one_month_after(start_date)

        hub_client = HubApiClient(hub_token, hub_url)
        ids_client = IdentityStoreApiClient(id_store_token, id_store_url)
        sbm_client = StageBasedMessagingApiClient(sbm_token, sbm_url)
        ms_client = MessageSenderApiClient(ms_token, ms_url)

        workbook = self.workbook_class()
        sheet = workbook.add_sheet('Registrations by date', 0)
        self.handle_registrations(sheet, hub_client, ids_client,
                                  start_date, end_date)

        sheet = workbook.add_sheet('Health worker registrations', 1)
        self.handle_health_worker_registrations(
            sheet, hub_client, ids_client, start_date, end_date)

        sheet = workbook.add_sheet('Enrollments', 2)
        self.handle_enrollments(sheet, sbm_client, ids_client, start_date,
                                end_date)

        sheet = workbook.add_sheet('SMS delivery per MSISDN', 3)
        self.handle_sms_delivery_msisdn(sheet, ms_client, start_date, end_date)

        sheet = workbook.add_sheet('Opt Outs by Subscription', 4)
        self.handle_optouts_by_subscription(
            sheet, sbm_client, ids_client, start_date, end_date)

        workbook.save(output_file)

        if email_recipients:
            file_name = 'report-%s-to-%s.xlsx' % (
                start_date.strftime('%Y-%m-%d'),
                end_date.strftime('%Y-%m-%d'))
            self.send_email(email_subject, file_name, output_file,
                            email_sender, email_recipients)

    def send_email(self, subject, file_name, file_location,
                   sender, recipients):
        email = EmailMessage(subject, '', sender, recipients)
        with open(file_location, 'rb') as fp:
            email.attach(file_name, fp.read(), 'application/vnd.ms-excel')
        email.send()

    def get_identity(self, ids_client, identity):
        if identity in self.identity_cache:
            return self.identity_cache[identity]

        identity_object = ids_client.get_identity(identity)
        self.identity_cache[identity] = identity_object
        return identity_object

    def get_messageset(self, sbm_client, messageset):
        if messageset in self.messageset_cache:
            return self.messageset_cache[messageset]

        messageset_object = sbm_client.get_messageset(messageset)
        self.messageset_cache[messageset] = messageset_object
        return messageset_object

    def get_registrations(self, hub_client, **kwargs):
        registrations = hub_client.get_registrations(kwargs)
        cursor = registrations['next']
        while cursor:
            for result in registrations['results']:
                yield result
            params = parse_cursor_params(cursor)
            registrations = hub_client.get_registrations(params)
            cursor = registrations['next']
        for result in registrations['results']:
            yield result

    def get_subscriptions(self, sbm_client, **kwargs):
        subscriptions = sbm_client.get_subscriptions(kwargs)
        cursor = subscriptions['next']
        while cursor:
            for result in subscriptions['results']:
                yield result
            params = parse_cursor_params(cursor)
            subscriptions = sbm_client.get_subscriptions(params)
            cursor = subscriptions['next']
        for result in subscriptions['results']:
            yield result

    def get_outbounds(self, ms_client, **kwargs):
        outbounds = ms_client.get_outbounds(kwargs)
        cursor = outbounds['next']
        while cursor:
            for result in outbounds['results']:
                yield result
            params = parse_cursor_params(cursor)
            outbounds = ms_client.get_outbounds(params)
            cursor = outbounds['next']
        for result in outbounds['results']:
            yield result

    def get_optouts(self, ids_client, **kwargs):
        optouts = ids_client.get_optouts(kwargs)
        cursor = optouts['next']
        while cursor:
            for result in optouts['results']:
                yield result
            params = parse_cursor_params(cursor)
            optouts = ids_client.get_optouts(params)
            cursor = optouts['next']
        for result in optouts['results']:
            yield result

    def handle_registrations(self, sheet, hub_client, ids_client,
                             start_date, end_date):

        sheet.set_header([
            'MSISDN',
            'Created',
            'gravida',
            'msg_type',
            'last_period_date',
            'language',
            'msg_receiver',
            'voice_days',
            'Voice_times',
            'preg_week',
            'reg_type',
            'Personnel_code',
            'Facility',
            'Cadre',
            'State',
        ])

        registrations = self.get_registrations(
            hub_client,
            created_after=start_date.isoformat(),
            created_before=end_date.isoformat())

        for idx, registration in enumerate(registrations):
            data = registration.get('data', {})
            operator_id = data.get('operator_id')
            receiver_id = data.get('receiver_id')

            if operator_id:
                operator_identity = self.get_identity(
                    ids_client, operator_id)
            else:
                operator_identity = {}

            if receiver_id:
                receiver_identity = self.get_identity(
                    ids_client, receiver_id)
            else:
                receiver_identity = {}

            operator_details = operator_identity.get('details', {})
            receiver_details = receiver_identity.get('details', {})
            default_addr_type = receiver_details.get('default_addr_type')
            if default_addr_type:
                addresses = receiver_details.get('addresses', {})
                msisdns = addresses.get(default_addr_type, {}).keys()
            else:
                msisdns = []

            sheet.add_row({
                'MSISDN': ','.join(msisdns),
                'Created': registration['created_at'],
                'gravida': data.get('gravida'),
                'msg_type': data.get('msg_type'),
                'last_period_date': data.get('last_period_date'),
                'language': data.get('language'),
                'msg_receiver': data.get('msg_receiver'),
                'voice_days': data.get('voice_days'),
                'Voice_times': data.get('voice_times'),
                'preg_week': data.get('preg_week'),
                'reg_type': data.get('reg_type'),
                'Personnel_code': operator_details.get('personnel_code'),
                'Facility': operator_details.get('facility_name'),
                'Cadre': operator_details.get('role'),
                'State': operator_details.get('state'),
            })

    def handle_health_worker_registrations(
            self, sheet, hub_client, ids_client, start_date, end_date):
        sheet.set_header([
            'Unique Personnel Code',
            'Facility',
            'State',
            'Cadre',
            'Number of Registrations'])

        registrations = self.get_registrations(
            hub_client,
            created_after=start_date.isoformat(),
            created_before=end_date.isoformat())

        registrations_per_operator = collections.defaultdict(int)

        for registration in registrations:
            operator_id = registration.get('data', {}).get('operator_id')
            registrations_per_operator[operator_id] += 1

        for operator_id, count in registrations_per_operator.items():
            operator = self.get_identity(ids_client, operator_id) or {}
            operator_details = operator.get('details', {})
            sheet.add_row({
                'Unique Personnel Code': operator_details.get(
                    'personnel_code'),
                'Facility': operator_details.get('facility_name'),
                'State': operator_details.get('state'),
                'Cadre': operator_details.get('role'),
                'Number of Registrations': count,
            })

    def handle_enrollments(self, sheet, sbm_client, ids_client, start_date,
                           end_date):

        sheet.set_header([
            'Message set',
            'Roleplayer',
            'Total enrolled',
            'Enrolled in period',
            'Enrolled and opted out in period',
            'Enrolled and completed in period',
        ])

        subscriptions = self.get_subscriptions(
            sbm_client,
            created_before=end_date.isoformat())

        data = collections.defaultdict(partial(collections.defaultdict, int))
        for subscription in subscriptions:
            messageset = self.get_messageset(
                            sbm_client, subscription['messageset'])
            identity = self.get_identity(ids_client, subscription['identity'])

            messageset_name = messageset['short_name'].split('.')[0]

            receiver_role = 'None'
            if identity:
                receiver_role = identity.get('details', {}).get(
                                    'receiver_role', 'None')

            data[messageset_name, receiver_role]['total'] += 1

            if parse_datetime(subscription['created_at']) > start_date:
                data[messageset_name, receiver_role]['total_period'] += 1

                if (not subscription['active'] and
                        not subscription['completed']):
                    data[messageset_name, receiver_role]['optouts'] += 1

                if subscription['completed']:
                    data[messageset_name, receiver_role]['completed'] += 1

        for key in sorted(data.keys()):
            sheet.add_row({
                1: key[0],
                2: key[1],
                3: data[key]['total'],
                4: data[key]['total_period'],
                5: data[key]['optouts'],
                6: data[key]['completed'],
            })

    def handle_sms_delivery_msisdn(
            self, sheet, ms_client, start_date, end_date):

        outbounds = self.get_outbounds(
            ms_client,
            created_after=start_date.isoformat(),
            created_before=end_date.isoformat()
        )

        data = collections.defaultdict(dict)
        count = collections.defaultdict(int)
        for outbound in outbounds:
            if 'voice_speech_url' not in outbound.get('metadata', {}):

                count[outbound['to_addr']] += 1
                data[outbound['to_addr']][outbound['created_at']] = \
                    outbound['delivered']

        if count != {}:
            max_col = max(count.values())

            header = ['MSISDN']
            for col_idx in range(0, max_col):
                header.append('SMS {}'.format(col_idx + 1))

            sheet.set_header(header)

            for msisdn, sms_data in data.items():

                row = {1: msisdn}

                for index, (key, state) in enumerate(sorted(sms_data.items())):
                    row[index+2] = 'Yes' if state else 'No'

                sheet.add_row(row)

    def handle_optouts_by_subscription(
            self, sheet, sbm_client, ids_client, start_date, end_date):

        sheet.set_header([
            "Timestamp",
            "Subscription Message Set",
            "Receiver's Role",
            "Reason",
        ])

        optouts = self.get_optouts(
            ids_client, created_at__gte=start_date.isoformat(),
            created_at__lte=end_date.isoformat())

        for optout in optouts:
            if 'identity' not in optout or not optout['identity']:
                message_set = "Unknown"
                receivers_role = "Unknown"
            else:
                identity = self.get_identity(ids_client, optout['identity'])
                receivers_role = identity.get('data', {}).get(
                    'role', 'Unknown')
                # Get the last subscription before the optout that is inactive
                subscriptions = list(self.get_subscriptions(
                    sbm_client, identity=optout['identity'],
                    created_before=optout['created_at'],
                    active=False, completed=False))
                subscriptions.sort(key=lambda s: s['created_at'], reverse=True)
                try:
                    subscription = subscriptions[0]
                except IndexError:
                    message_set = "Unknown"
                else:
                    message_set = self.get_messageset(
                        sbm_client, subscription['messageset'])['short_name']

            sheet.add_row({
                "Timestamp": optout['created_at'],
                "Subscription Message Set": message_set,
                "Receiver's Role": receivers_role,
                "Reason": optout['reason'],
            })

        # Add a warning to the sheet, because we cannot guarantee that the
        # subscription that we choose is the subscription that was opted out of
        sheet.add_row({
            "Timestamp": (
                "NOTE: The message set is not guaranteed to be correct, as "
                "the current structure of the data does not allow us to link "
                "the opt out to a subscription, so this is a best-effort "
                "guess."),
        })
